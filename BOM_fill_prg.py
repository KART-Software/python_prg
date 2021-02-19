# 2020年度のフォーマットで正しく動作

import pathlib  # 標準ライブラリ
import openpyxl  # 外部ライブラリ(ダウンロードが必要)
import os  # 標準ライブラリ

def sub_total_reader(content : str , bom_stem : int , fca_stem : int ):


    """
    FCAからsub_totalを読み取り，BOMに移す関数

    B列を1から順に見ていき，contentと一致するものを探す．
    contentとの一致部から下に見ていき，初めて空欄に達した行の
    fca_stemの列にsub_totalの値が入っている．
    それをbomの指定の箇所に転記する
    """

    
    for j in range(1,500):
        if sh.cell(j, 2).value == content :
            for k in range(j,500):
                if sh.cell(k,2).value == None :
                    sheet_BOM.cell(i,bom_stem).value = float(sh.cell(k,fca_stem).value)
                    break
            # print(sheet_name + content + 'あり') -> 全てのシートが開かれたか確認できる
            break

# 注意：python_prgのフォルダをコストレポートの最上階層と同じ階層に置くこと

print('本プログラムはシステムアセンブリごとに動かしてください')
path_CR_input = str(input('最上階層のフォルダ名：'))
path_CR = pathlib.Path(os.path.join('..', path_CR_input))
# 最上階層のフォルダ名を入力

path_BOM_input = str(input('BOMのファイル名を入力(拡張子も含める)：'))
path_BOM = pathlib.Path(os.path.join(path_CR, path_BOM_input))
# BOMのパスを入力

car_number = str(input('車番(半角数字三桁)：'))
chapter = int(input('何章ですか(半角数字)：'))
path_SA_input = str(input('システムアセンブリのフォルダ名を入力：'))
path_SA = pathlib.Path(os.path.join(path_CR, path_SA_input))
# システムアセンブリのパスを入力

try:
    BOM = openpyxl.load_workbook(path_BOM)
    sheet_BOM = BOM['BOM']
except:
    print('BOMが開けません')
# BOMファイルを開く(シート名がBOMになっていることを確認)

print('BOMファイル内で指定したシステムアセンブリが記入される行の最小値と最大値を入力してください')
min_row = int(input('最小値：'))
max_row = int(input('最大値：'))

counter_file = 0
for sub_dir in list(path_SA.glob('*')):  # システムアセンブリ内のすべてのサブディレクトリをリスト化
    for FCA_files in sub_dir.glob('*.xlsx'):  # サブディレクトリ内のすべてのエクセルファイルを順に開く
        counter_file += 1  # 開くFCAのファイル数をカウント
        fca = openpyxl.load_workbook(FCA_files,data_only=True)  # 値のみの読み取り
        sheet_name_list = fca.get_sheet_names()
        for x, sh in enumerate(fca):              # 開いたエクセルファイル内の全シートを順に読み込む
            sheet_name_space =  sheet_name_list[x]   # シート名を取得
            sheet_name = sheet_name_space.replace(' ','')  # スペースを除いた上でシート名を確認
            print(sheet_name)
            for i in range(min_row, max_row + 1):
                BOM_component_space =  sheet_BOM.cell(i,6).value  # Quantityを埋める
                BOM_component = BOM_component_space.replace(' ','')
                if sheet_name == BOM_component:  # F列にパーツ名が入るものとして動作
                    sheet_BOM.cell(i, 9).value = int(sh.cell(2,14).value)
                    sub_total_reader('Material', 10 , 14)
                    sub_total_reader('Process', 11 , 9)
                    sub_total_reader('Fastener', 12 , 10)
                    sub_total_reader('Tooling', 13 , 9)   
                    
                    sheet_BOM.cell(i,15).value = '=HYPERLINK("[..\\' + car_number + \
                    '_KyotoUniversity_FSAEJ_CR\\' + path_SA_input + '\\' + \
                    car_number + '_A' + str(chapter) + '-' + str(counter_file) + '_' +  sheet_BOM.cell(i,2).value + \
                    '\\' + car_number + '_KyotoUniversity_FSAEJ_CR_FCA_' + sheet_BOM.cell(i,2).value + \
                    '.xlsx]\'' + sheet_name_space + '\'!A1",F' + str(i) + ')'

                    # HYPERLINK関数を文字列の処理のみで使う

for t in range(min_row, max_row + 1):
    if 'A' in str(sheet_BOM.cell(t,3).value) and str(sheet_BOM.cell(t,9).value) != '1':
        for u in range(t+1,500):
            if sheet_BOM.cell(u,5).value != None :
               sheet_BOM.cell(u,9).value *= sheet_BOM.cell(t,9).value
            elif sheet_BOM.cell(u,5).value == None :
                break

                # アセンブリが複数セットある場合に，属する部品数にもセット数を掛けておく
    
print('総ファイル数',counter_file)
BOM.save(path_BOM)
