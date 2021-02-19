"""
一行当たりの文字数がPEP8を無視しています．申し訳ないです．
読みにくかったら，改行してください．
PEP8では一行当たりの文字数は79文字以下が推奨されています．
また，インデントもTabキーを使用しています．
(本モジュールはVisual Studioで編集しています．)
"""

import openpyxl  # 外部ライブラリ
import os
import pathlib

print('コストレポートの最上階層と同じ階層にpython_prgのフォルダとCost_Tablesのフォルダをコピーして下さい．')

path_CR_input = str(input('コストレポートの最上階層のフォルダ名を入力:'))
path_SS_input = str(input('実行するシステムアセンブリのフォルダ名を入力:'))
path_SS = pathlib.Path(os.path.join('..' ,path_CR_input, path_SS_input))
# システムアセンブリのパスを生成

MA_CT_name = str(input('Materialのコストテーブルのファイル名を入力(拡張子あり)：'))
PR_CT_name = str(input('Processのコストテーブルのファイル名を入力(拡張子あり)：'))
FA_CT_name = str(input('Fastenerのコストテーブルのファイル名を入力(拡張子あり)：'))
TO_CT_name = str(input('Toolingのコストテーブルのファイル名を入力(拡張子あり)：'))
MP_CT_name = str(input('Multiplierのコストテーブルのファイル名を入力(拡張子あり)：'))
# コストテーブルのファイル名を入力

path_MA_CT = pathlib.Path(os.path.join('..\Cost_Tables' ,MA_CT_name))
path_PR_CT = pathlib.Path(os.path.join('..\Cost_Tables' ,PR_CT_name))
path_FA_CT = pathlib.Path(os.path.join('..\Cost_Tables' ,FA_CT_name))
path_TO_CT = pathlib.Path(os.path.join('..\Cost_Tables' ,TO_CT_name))
path_MP_CT = pathlib.Path(os.path.join('..\Cost_Tables' ,MP_CT_name))
# コストテーブルのパスを生成

MA_CT = openpyxl.load_workbook(path_MA_CT)
PR_CT = openpyxl.load_workbook(path_PR_CT)
FA_CT = openpyxl.load_workbook(path_FA_CT)
TO_CT = openpyxl.load_workbook(path_TO_CT)
MP_CT = openpyxl.load_workbook(path_MP_CT)
# コストテーブルを開く

MA_CT_sh = MA_CT.active
PR_CT_sh = PR_CT.active
FA_CT_sh = FA_CT.active
TO_CT_sh = TO_CT.active
MP_CT_sh = MP_CT.active
# アクティブシートを選択

for folder_obj in path_SS.iterdir():
    for pass_obj in folder_obj.iterdir():
    # システムアセンブリフォルダ内のファイルをワイルドカード検索
        if pass_obj.match('*.xlsx'):
            fca = openpyxl.load_workbook(pass_obj)
            sheet_name_list = fca.get_sheet_names()
            # get_sheet_names関数は使えなくなる可能性あり
            for m, sh in enumerate(fca):
                for i in range(1, 500):
                    if sh['B' + str(i)].value == 'Material':
                        for j in range(i+1, 500):
                            if sh['B' + str(j)].value == 'None':
                                break
                            # 'None'の文字列がヒットした場合にbreak
                            if sh['B' + str(j)].value == None and sh['A' + str(j)].value == None:
                                break
                            # A列，B列共に空欄に達した場合にbreak
                            else:
                                material_name_space = str(sh['B'+str(j)].value)
                                material_name = material_name_space.replace(' ', '')
                                # スペースを除去
                                for k in range(3, 1501):
                                    MA_CT_index_space = str(MA_CT_sh['B'+str(k)].value)
                                    MA_CT_index = MA_CT_index_space.replace(' ','')
                                    # コストテーブルのインデックスからスペースを除去
                                    if material_name == MA_CT_index and isinstance(MA_CT_sh['E'+str(k)].value, str):
                                        break
                                    """
                                    FCAのMaterial名とコストテーブルのインデックスが一致するも，
                                    コストテーブル側に数値ではなく，数式が入っている場合にbreak
                                    本モジュールではUnitCostが数式の場合は処理できません
                                    """
                                    if material_name == MA_CT_index and MA_CT_sh['E'+str(k)].value == sh['D' + str(j)].value:
                                        break
                                    # FCAのMaterial名とコストテーブルのインデックスが一致し，UnitCostが正しい場合break
                                    elif material_name == MA_CT_index and MA_CT_sh['E'+str(k)].value != sh['D' + str(j)].value:
                                        print(sheet_name_list[m] + ' : ' + material_name + 'のUnitCostが誤っています.')
                                        break
                                    # FCAのMaterial名とコストテーブルのインデックスが一致し，UnitCostが誤っている場合，誤りを表示
                                    if k == 1500:
                                        print(sheet_name_list[m] + ' : ' + material_name + 'のMaterial名が誤っています.')
                                    # コストテーブル内にFCAのMaterialと一致するものが見つけられなかった場合, 誤りを表示
                        break

                # 以下，Materialと同内容のコード, 関数にすると引数が多くて面倒だったので繰り返しになっています
                for i in range(1, 500):
                    if sh['B' + str(i)].value == 'Process':
                        for j in range(i+1, 500):
                            if sh['B' + str(j)].value == 'None':
                                break
                            if sh['B' + str(j)].value == None and sh['A' + str(j)].value == None:
                                break
                            else:
                                process_name_space = str(sh['B'+str(j)].value)
                                process_name = process_name_space.replace(' ', '')
                                for k in range(3, 201):
                                    PR_CT_index_space = str(PR_CT_sh['B'+str(k)].value)
                                    PR_CT_index = PR_CT_index_space.replace(' ','') 
                                    if process_name == PR_CT_index and isinstance(PR_CT_sh['C'+str(k)].value, str):
                                        break
                                    if process_name == PR_CT_index and PR_CT_sh['C'+str(k)].value == sh['D' + str(j)].value:
                                        if PR_CT_sh['D'+str(k)].value != sh['E' + str(j)].value:
                                            print(sheet_name_list[m] + ' : ' + process_name + 'のUnitが誤っています.')
                                        break
                                    elif process_name == PR_CT_index and PR_CT_sh['C'+str(k)].value != sh['D' + str(j)].value:
                                        print(sheet_name_list[m] + ' : ' + process_name + 'のUnitCostが誤っています.')
                                        if PR_CT_sh['D'+str(k)].value != sh['E' + str(j)].value:
                                            print(sheet_name_list[m] + ' : ' + process_name + 'のUnitが誤っています.')
                                        break
                                    if k == 200:
                                        print(sheet_name_list[m] + ' : ' + process_name + 'のProcess名が誤っています.')
                        break
                
                for i in range(1, 500):
                    if sh['B' + str(i)].value == 'Fastener':
                        for j in range(i+1, 500):
                            if sh['B' + str(j)].value == 'None':
                                break
                            if sh['B' + str(j)].value == None and sh['A' + str(j)].value == None:
                                break
                            else:
                                fastener_name_space = str(sh['B'+str(j)].value)
                                fastener_name = fastener_name_space.replace(' ', '')
                                for k in range(3, 101):
                                    FA_CT_index_space = str(FA_CT_sh['B'+str(k)].value)
                                    FA_CT_index = FA_CT_index_space.replace(' ','') 
                                    if fastener_name == FA_CT_index and isinstance(FA_CT_sh['D'+str(k)].value, str):
                                        break
                                    if fastener_name == FA_CT_index and FA_CT_sh['D'+str(k)].value == sh['D' + str(j)].value:
                                        break
                                    elif fastener_name == FA_CT_index and FA_CT_sh['D'+str(k)].value != sh['D' + str(j)].value:
                                        print(sheet_name_list[m] + ' : ' + fastener_name + 'のUnitCostが誤っています.')
                                        break
                                    if k == 100:
                                        print(sheet_name_list[m] + ' : ' + fastener_name + 'のfastener名が誤っています.')
                        break

                for i in range(1, 500):
                    if sh['B' + str(i)].value == 'Tooling':
                        for j in range(i+1, 500):
                            if sh['B' + str(j)].value == 'None':
                                break
                            if sh['B' + str(j)].value == None and sh['A' + str(j)].value == None:
                                break
                            else:
                                tooling_name_space = str(sh['B'+str(j)].value)
                                tooling_name = tooling_name_space.replace(' ', '')
                                for k in range(3, 51):
                                    TO_CT_index_space = str(TO_CT_sh['C'+str(k)].value)
                                    TO_CT_index = TO_CT_index_space.replace(' ','') 
                                    if tooling_name == TO_CT_index and isinstance(TO_CT_sh['D'+str(k)].value, str):
                                        break
                                    if tooling_name == TO_CT_index and TO_CT_sh['D'+str(k)].value == sh['D' + str(j)].value:
                                        break
                                    elif tooling_name == TO_CT_index and TO_CT_sh['D'+str(k)].value != sh['D' + str(j)].value:
                                        print(sheet_name_list[m] + ' : ' + tooling_name + 'のUnitCostが誤っています.')
                                        break
                                    if k == 50:
                                        print(sheet_name_list[m] + ' : ' + tooling_name + 'のtooling名が誤っています.')
                        break

                for i in range(1, 500):
                    if sh['G' + str(i)].value == 'Multiplier':
                        for j in range(i+1, 500):
                            if sh['G' + str(j)].value == None and sh['A' + str(j)].value == None:
                                break
                            elif sh['G' + str(j)].value == None:
                                continue
                            elif 'Repeat' in str(sh['G' + str(j)].value):
                                continue
                            else:
                                multiplier_name_space = str(sh['G'+str(j)].value)
                                multiplier_name = multiplier_name_space.replace(' ', '')
                                for k in range(3, 51):
                                    MP_CT_index_space = str(MP_CT_sh['B'+str(k)].value)
                                    MP_CT_index = MP_CT_index_space.replace(' ','') 
                                    if multiplier_name in MP_CT_index and isinstance(MP_CT_sh['C'+str(k)].value, str):
                                        break
                                    if multiplier_name in MP_CT_index and MP_CT_sh['C'+str(k)].value == sh['H' + str(j)].value:
                                        break
                                    elif multiplier_name in MP_CT_index and MP_CT_sh['C'+str(k)].value != sh['H' + str(j)].value:
                                        print(sheet_name_list[m] + ' : ' + multiplier_name + 'のMultiVal.が誤っています.')
                                        break
                                    if k == 50:
                                        print(sheet_name_list[m] + ' : ' + multiplier_name + 'のMultiplier名が誤っています.')
                        break                   
